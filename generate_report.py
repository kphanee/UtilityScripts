#!/usr/bin/python

from datetime import date
from datetime import timedelta
from lxml import html

import json
import pprint
import re
import requests
import subprocess
import sys

try:
    from openpyxl.workbook import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side, Style
    from openpyxl.cell import Cell
    from openpyxl.cell import get_column_letter
except ImportError:
    print('You need to execute:\n\tsudo pip install openpyxl')
    sys.exit(-1)

repositories = [
    {
        'name': 'blink',
        'path': '/home/kphanee/workspace/oss-report-generation/blink',
        'type': 'git',
        'host': 'https://chromium.googlesource.com/',
        'guid': '82804f94bf4b4d551c864ff2ef44317633f76246'
    },
    {
        'name': 'chromium',
        'path': '/home/kphanee/workspace/oss-report-generation/chromium',
        'type': 'git',
        'host': 'https://chromium.googlesource.com/',
        'guid': 'c14d891d44f0afff64e56ed7c9702df1d807b1ee'
    },
    {
        'name': 'trace-viewer',
        'path': '/home/kphanee/workspace/oss-report-generation/trace-viewer',
        'type': 'git',
        'host': 'https://github.com',
        'url': 'https://github.com/google/trace-viewer'
    }
    # {
    #     'name': 'skia',
    #     'path': '/home/kphanee/workspace/oss-report-generation/skia',
    #     'type': 'git',
    #     'host': 'https://chromium.googlesource.com/',
    #     'guid': '586101c79b0490b50623e76c71a5fd67d8d92b08'
    # },
    # {
    #     'name': 'v8',
    #     'path': '/home/kphanee/workspace/oss-report-generation/v8',
    #     'type': 'git',
    #     'host': 'https://chromium.googlesource.com/',
    #     'guid': '33f2fb0e53d135f0ee17cfccd9d993eb2a6f47de'
    # }
]

authors = [
    { 'name': 'Abhijeet Kandalkar', 'email': 'abhijeet.k@samsung.com' },
    { 'name': 'Ajay Berwal', 'email': 'ajay.berwal@samsung.com' },
    { 'name': 'Akhil Teeka Dhananjaya', 'email': 'akhil.td@samsung.com' },
    { 'name': 'Behara Mani Shyam Patro', 'email': 'behara.ms@samsung.com' },
    { 'name': 'Gandhi Kishor Addanki', 'email': 'kishor.ag@samsung.com' },
    { 'name': 'Ganesh Kamat', 'email': 'ganesh.kamat@samsung.com' },
    { 'name': 'Kaja Mohaideen', 'email': 'kaja.m@samsung.com' },
    { 'name': 'Karthik Gopalan', 'email': 'karthikg.g@samsung.com' },
    { 'name': 'Kulajit Das', 'email': 'das.kulajit@samsung.com' },
    { 'name': 'Mallikarjuna Narala', 'email': 'mallik.n@samsung.com' },
    { 'name': 'Munukutla Subrahmanya Praveen', 'email': 'sataya.m@samsung.com' },
    { 'name': 'Nikhil Sahni', 'email': 'nikhil.sahni@samsung.com' },
    { 'name': 'Pavan Kumar Emani', 'email': 'pavan.e@samsung.com' },
    { 'name': 'Prabhavathi Perumal', 'email': 'prabha.p@samsung.com' },
    { 'name': 'Prashant Nevase', 'email': 'prashant.n@samsung.com' },
    { 'name': 'Putturaju R', 'email': 'puttaraju.r@samsung.com' },
    { 'name': 'Ravi Kasibhatla', 'email': 'r.kasibhatla@samsung.com', 'username': 'kphanee' },
    { 'name': 'Shanmuga Pandi', 'email': 'shanmuga.m@samsung.com' },
    { 'name': 'Siva Gunturi', 'email': 'siva.gunturi@samsung.com' },
    { 'name': 'Sohan Jyoti Ghosh', 'email': 'sohan.jyoti@samsung.com' },
    { 'name': 'Suchit Agarwal', 'email': 'a.suchit@samsung.com' },
    { 'name': 'Sujith S S', 'email': 'sujiths.s@samsung.com' },
    { 'name': 'Suyash Sengar', 'email': 'suyash.s@samsung.com' },
    { 'name': 'Tanvir Rizvi', 'email': 'tanvir.rizvi@samsung.com' },
    { 'name': 'Thanikassalam Kankayan', 'email': 'thanik.k@samsung.com' },
    { 'name': 'Vivek Agrawal', 'email': 'vivek.s14@samsung.com' },
    { 'name': 'Vivek Galatage', 'email': ['vivek.vg@samsung.com', 'vivekgalatage@gmail.com'] }
]

def execute(cwd, command, verbose=False, progress=False):
    process = subprocess.Popen(command, shell=True, cwd=cwd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    processOutput = []
    while True:
        nextline = process.stdout.readline()
        if nextline == '' and process.poll() != None:
            break
        processOutput.append(nextline)
        if verbose:
            helper.writeln(nextline)
        if progress:
            helper.writeln('.')
    if progress:
        helper.writeln('')
    output = process.communicate()[0]
    exitCode = process.returncode

    if exitCode == 0:
        return processOutput
    else:
        raise ProcessException(command, exitCode, output)

def updateRepositories():
    for repo in repositories:
        helper.writeln('Updating repository: %s' % repo['name'], '\n')
        if (repo['type'] == 'git'):
            output = execute(repo['path'], 'git pull', progress=True)
    helper.writeln('')


"""
Utility class with some of the common functions for general usage.
"""
class UtilityHelper(object):
    def __init__(self, authorColWidth, startingYear=2013):
        self._previous_text = ''
        self._max_width = { 'authorColWidth': authorColWidth }
        self._duration = { 'total': 0, 'year': 1, 'week': 2 }
        self._current_week = None
        self._recording_duration = range(startingYear, date.today().year + 1)

    def __getattr__(self, attr):
        if attr in self._max_width:
            return self._max_width[attr]
        elif attr in self._duration:
            return self._duration[attr]

    def writeln(self, text, newline=None):
        if not newline:
            sys.stdout.write(text)
        elif newline == '\r' or text == '':
            sys.stdout.write('\n')
        elif text != self._previous_text:
            sys.stdout.write(text)
            self._previous_text = text
        sys.stdout.flush()

    def GetFormattedAuthorEmails(self):
        authorEmails = []
        for author in authors:
            self._max_width['authorColWidth'] = max(self._max_width['authorColWidth'], len(author['name']))
            if (type(author['email']) == list):
                authorEmails = authorEmails +  author['email']
            elif (type(author['email']) == str):
                authorEmails.append(author['email'])

        authorsList = []
        for email in authorEmails:
            authorsList.append('\(' + email + '\)')
        return '\|'.join(authorsList)

    def GetAuthorColumnWidth(self):
        return self._max_width['authorColWidth']

    def GetContributionsReportingYearRange(self):
        return self._recording_duration

    def GetCurrentWeek(self):
        if self._current_week is None:
            day = date.today() - timedelta(days=7)
            start = day - timedelta(days=day.weekday())
            end = start + timedelta(days=6)
            self._current_week = (start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d'))
        return self._current_week

    def GetWeekNumberFromDate(self, week=None, sep='-'):
        if not week:
            week = self._current_week[0]
        weekStart = week.split(sep)
        assert(len(weekStart) == 3)
        return date(weekStart[0], weekStart[1], weekStart[2]).isocalendar()[1]

    def GetDatesFromWeekNumber(self, year, week):
        start = date(year, 1, 1)
        if(start.weekday() > 3):
            start = start + timedelta(7 - start.weekday())
        else:
            start = start - timedelta(start.weekday())
        start = start + timedelta(days=(week)*7)
        end = start + timedelta(days=6)
        return (start.strftime('%Y-%m-%d'), end.strftime('%Y-%m-%d'))

"""
Generates the .xlsx report with formatted data in report view.
"""
class ExcelWorkbook(object):
    def __init__(self):
        self.__thin_border = Border(left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))
        self.__greyTheme = Style(border=self.__thin_border,
                                 fill=PatternFill(fill_type='solid',
                                                  start_color='FFD8D8D8'))
        self.__whiteTheme = Style(border=self.__thin_border,
                                  fill=PatternFill(fill_type='solid',
                                                   start_color='FFFFFFFF'))
        self._repoToColumnTuples = [('B', 'chromium'),
                                    ('C', 'blink'),
                                    ('D', 'trace-viewer'),
                                    ('E', 'skia'),
                                    ('F', 'v8'),
                                    ('G', 'Total')]
        self._workbook = Workbook()
        self._author_col_width = helper.GetAuthorColumnWidth()
        self._data_col_width = 0
        self._header_row_height = 30
        self._author_data = None
        self._current_sheet = None

    def column_number_to_letter(self, col):
        return get_column_letter(col)

    def GenerateCompleteContributionsReport(self):
        self._ReadContributionsDataFromJson()
        self._GenerateTotalContributionSheet()
        self._GenerateYearlyContributionSheet()
        self._GenerateWeeklyContributionSheet(str(date.today().year))

        self._workbook.save('weeklyReport.xlsx')

    def _ReadContributionsDataFromJson(self):
        with open('weeklyReport.json') as jsonFile:
            self._author_data = json.loads(jsonFile.read())

    def _NextSheetIndex(self):
        return self._workbook.get_index(self._workbook.get_sheet_by_name('Sheet'))

    def _GenerateTotalContributionSheet(self):
        self._current_sheet = self._workbook.create_sheet(index=self._NextSheetIndex(), title='Total')
        self._PopulateContributions('total')
        self._CalculateContributionsSummation()
        self._BeautifyWorksheet()

    def _GenerateYearlyContributionSheet(self):
        for year in helper.GetContributionsReportingYearRange():
            self._current_sheet = self._workbook.create_sheet(index=self._NextSheetIndex(),
                                                              title=str(year) + ' Contributions')
            self._PopulateContributions('total', str(year))
            self._CalculateContributionsSummation()
            self._BeautifyWorksheet()

    def _GenerateWeeklyContributionSheet(self, year):
        for repo in repositories:
            repoName = repo['name']
            self._current_sheet = self._workbook.create_sheet(index=self._NextSheetIndex(),
                                                              title='%s %s Weekly' % (year, repoName))
            for week in xrange(52, 0, -1):
                (begin, end) = helper.GetDatesFromWeekNumber(int(year), week)
                key = 'W%02d %s to %s' % (week, begin, end)
                self._PopulateContributions(key, year, repoName)

            self._CalculateContributionsSummation()
            self._BeautifyWorksheet()

    def _GenerateWeeklyClosedContributionsSheet(self, year):
        # Patches which got closed during the week.
        self._current_sheet = self._workbook.create_sheet(index=self._NextSheetIndex(),
                                                          title='Weekly closed contributions')
        self._PopulateWeeklyClosedContributionsDetails()
        self._BeautifyWorksheet()

    def _GenerateWeeklyOpenContributionsSheet(self, year):
        # Patches which are under review during the week.
        self._current_sheet = self._workbook.create_sheet(index=self._NextSheetIndex(),
                                                          title='Weekly open contributions')
        self._PopulateWeeklyOpenContributionsDetails()
        self._BeautifyWorksheet()

    def _PopulateWeeklyClosedContributionsSheet(self):
        currentWeek = helper.GetCurrentWeek()
        key = currentWeek[0] + ' to ' + currentWeek[1]
        self._current_sheet['A1'] = 'Name'
        # self._current_sheet['A2']

    def _PopulateContributions(self, key, year=None, repo=None):
        codeReviewSearchURL = """https://codereview.chromium.org/search?closed=1&owner=%s&reviewer=&cc=&repo_guid=&base=&project=&private=1&commit=1&created_before=&created_after=&modified_before=&modified_after=&order=&format=html&keys_only=False&with_messages=False&cursor=&limit=30"""

        self._CreateSheetHeaderRow(key)
        for i in xrange(len(self._author_data)):
            author = self._author_data[i]
            index = str(i + 2)
            self._current_sheet['A' + index] = author['name']

            # TODO: Correct the hyperlink as per sheet being filled.
            if type(author['email']) == list:
                self._current_sheet['A' + index].hyperlink = codeReviewSearchURL % author['email'][0]
            else:
                self._current_sheet['A' + index].hyperlink = codeReviewSearchURL % author['email']

            if 'contributions' not in author:
                continue

            contributions = author['contributions']
            if key.startswith('W'):
                self._PopulateWeeklyContributions(contributions, index, key, year, repo)
            else:
                self._PopulateYearlyContributions(contributions, index, key, year)

    def _PopulateWeeklyContributions(self, contributions, rowIndex, key, year, repo):
        assert(year is not None and repo is not None)
        column = self.column_number_to_letter(self._current_sheet.get_highest_column())
        if repo in contributions:
            if year in contributions[repo] and key in contributions[repo][year]:
                self._current_sheet[column + rowIndex] = contributions[repo][year][key]
            else:
                self._current_sheet[column + rowIndex] = 0
        else:
                self._current_sheet[column + rowIndex] = 0

    def _PopulateYearlyContributions(self, contributions, rowIndex, key, year=None):
        for (col, repo) in self._repoToColumnTuples:
            if repo in contributions:
                if year:
                    if year in contributions[repo]:
                        self._current_sheet[col + rowIndex] = contributions[repo][year][key]
                    else:
                        self._current_sheet[col + rowIndex] = 0
                else:
                    if key in contributions[repo]:
                        self._current_sheet[col + rowIndex] = contributions[repo][key]
                    else:
                        self._current_sheet[col + rowIndex] = 0
            else:
                self._current_sheet[col + rowIndex] = 0

        self._current_sheet['G' + rowIndex] = '=sum(b%s:f%s)' % (rowIndex, rowIndex)

    def _CalculateContributionsSummation(self):
        totalAuthors = len(self._author_data)
        finalRowIndex = str(totalAuthors + 2)
        dataRange = '(%s2:%s' + str(totalAuthors + 1) + ')'
        self._current_sheet['A' + finalRowIndex] = 'Total'
        for col in xrange(2, self._current_sheet.get_highest_column() + 1):
            column = self.column_number_to_letter(col)
            self._current_sheet[column + finalRowIndex] = '=sum' + (dataRange % (column, column))

    def _BeautifyWorksheet(self):
        sheet = self._current_sheet

        sheet.column_dimensions['A'].width = self._author_col_width
        sheet.row_dimensions[1].height = self._header_row_height
        sheet.row_dimensions[sheet.get_highest_row()].height = self._header_row_height

        lastRow = str(sheet.get_highest_row())
        header = [ 'A1' ]
        footer = [ 'A' + lastRow ]
        for col in xrange(2, sheet.get_highest_column() + 1):
            column = self.column_number_to_letter(col)
            sheet.column_dimensions[column].width = self._data_col_width
            header.append(column + '1')
            footer.append(column + lastRow)

        for col in header:
            sheet[col].style = Style(alignment=Alignment(vertical='bottom', horizontal='center'),
                                     border=self.__thin_border,
                                     fill=PatternFill(fill_type='solid', start_color='FFFFFF99'),
                                     font=Font(bold=True))

        for col in footer:
            sheet[col].style = Style(alignment=Alignment(vertical='bottom', horizontal='center'),
                                     border=self.__thin_border,
                                     fill=PatternFill(fill_type='solid', start_color='FFFFFF99'),
                                     font=Font(bold=True))

        for col in xrange(1, sheet.get_highest_column() + 1):
            for row in xrange(2, sheet.get_highest_row() - 1):
                column = self.column_number_to_letter(col)
                cell = sheet[str(column + str(row))]
                if (row % 2) == 0:
                    cell.style = self.__greyTheme
                else:
                    cell.style = self.__whiteTheme

    def _CreateSheetHeaderRow(self, key):
        subkey = key.split(' ', 1)
        self._current_sheet['A1'] = 'Name'
        if subkey[0].startswith('W'):
            column = self.column_number_to_letter(self._current_sheet.get_highest_column() + 1) + '1'
            self._current_sheet[column] = subkey[0]
            self._current_sheet[column].comment = Comment(subkey[1], 'OSS')
            # TODO: Better way to determine width of weekly column?
            self._data_col_width = 5
        else:
            for (col, repo) in self._repoToColumnTuples:
                self._current_sheet[col + '1'] = repo.title()
                self._data_col_width = max(self._data_col_width, len(repo) + 1)


"""
Generates the .json file which has all the details
"""
class AuthorContributions(object):
    def __init__(self, authorEmails):
        self._author_emails = authorEmails
        self._current_week = helper.GetCurrentWeek()

    def PopulateContributionsAndFormatToJson(self):
        self._PopulateAllContributions(helper.total)
        reportingDuration = helper.GetContributionsReportingYearRange() 
        for year in reportingDuration:
            self._PopulateAllContributions(helper.year, year=str(year))
            for week in xrange(1, 53):
                self._PopulateAllContributions(helper.week, str(year), week)
        self._PopulateLastWeekContributionDetails()

        with open('weeklyReport.json', 'w') as jsonFile:
            json.dump(authors, jsonFile, sort_keys=True, indent=4)

    def _PopulateAllContributions(self, duration, year=None, week=None):
        QUERY = [
            ('total', 'git shortlog -es --author="%s"'),
            ('Y%s', 'git shortlog -es --author="%s" --after="%s-01-01T00:00:00+00:00" --before="%s-12-31T23:59:59+00:00"'),
            ('W%02d %s to %s', 'git shortlog -es --author="%s" --after="%sT00:00:00+00:00" --before="%s-12-31T23:59:59+00:00"')
        ]

        for repo in repositories:
            key = ''
            command = ''
            if duration == helper.total:
                helper.writeln('Calculating total contributions so far, a big task you see!')
                key = QUERY[helper.total][0]
                command = QUERY[helper.total][1] % (self._author_emails)
            elif duration == helper.year:
                assert(year != None)
                helper.writeln('Figuring out what kept you busy in %s' % year)
                key = QUERY[helper.year][0] % year
                command = QUERY[helper.year][1] % (self._author_emails, year, year)
            elif duration == helper.week:
                assert(year != None and week != None)
                helper.writeln('Lets check out what are you upto currently in every week of %s' % (year))
                (begin, end) = helper.GetDatesFromWeekNumber(int(year), week)
                key = QUERY[helper.week][0] % (week, begin, end)
                command = QUERY[helper.week][1] % (self._author_emails, begin, end)
            helper.writeln('.', '\r')
            output = execute(repo['path'], command, progress=True)
            self._ParseGitLogOutput(repo, output, key, year)

    def _ParseGitLogOutput(self, repo, output, key, year=None):
        allContributions = {}
        for line in output:
            if (not len(line)):
                continue
            details = line.split()
            commits = int(details[0])
            retriedEmail = details[len(details) - 1]
            email = re.compile('((^<)(\w+\.?\w+@\w+.?\w+)(.*$))').match(retriedEmail).groups()[2]
            if email not in allContributions:
                allContributions[email] = commits
            else:
                allContributions[email] = allContributions[email] + commits
        if not len(allContributions):
            return

        for i in xrange(len(authors)):
            emailType = type(authors[i]['email'])
            commits = 0
            if emailType == list:
                for eachEmail in authors[i]['email']:
                    if eachEmail in allContributions:
                        commits = commits + allContributions[eachEmail]
            elif emailType == str:
                email = authors[i]['email']
                if email in allContributions:
                    commits = allContributions[email]
            if commits:
                repoName = repo['name']
                if 'contributions' not in authors[i]:
                    authors[i]['contributions'] = {}
                if repoName not in authors[i]['contributions']:
                    authors[i]['contributions'][repoName] = {}
                if year:
                    if year not in authors[i]['contributions'][repoName]:
                        authors[i]['contributions'][repoName][year] = {}
                    if key.startswith('Y'):
                        key = 'total'
                    authors[i]['contributions'][repoName][year][key] = commits
                else:
                    authors[i]['contributions'][repoName][key] = commits

    def _PopulateLastWeekContributionDetails(self):
        helper.writeln('Preparing the contribution details, hang on!')
        for repo in repositories:
            repoName = repo['name']
            for i in xrange(len(authors)):
                author = authors[i]
                helper.writeln('.')
                if 'contributions' not in authors[i]:
                    continue

                if repoName not in authors[i]['contributions']:
                    continue

                currentYear = date.today().year
                currentWeekNumber = helper.GetWeekNumberFromDate(self._current_week[0])
                currentWeekKey = 'W' + str(currentWeekNumber) + ' ' +
                                 self._current_week[0] + ' to ' + self._current_week[1]
                print currentYear, '->', currentWeekKey
                if currentYear not in authors[i]['contributions'][repoName] or
                   currentWeekKey not in authors[i]['contributions'][repoName][currentYear]:
                   continue

                assert('host' in repo)
                if repo['host'] == 'https://chromium.googlesource.com/':
                    assert(repo['guid'] != None)
                    issues = self._ProcessRietveld(author, repo['guid'])
                    closedIssues = issues[0]
                    openIssues = issues[1]
                elif repo['host'] == 'https://github.com':
                    assert(repo['url'] != None)
                    if 'username' in author:
                        issues = self._ProcessGithub(author, repo['url'])
                        closedIssues = issues[0]
                        openIssues = issues[1]

                if 'closed' not in authors[i]['contributions'][repoName]:
                    authors[i]['contributions'][repoName]['closed'] = []
                if 'open' not in authors[i]['contributions'][repoName]:
                    authors[i]['contributions'][repoName]['open'] = []

                for issue in closedIssues:
                    newIssue = {};
                    newIssue[issue[0]] = issue[1]
                    authors[i]['contributions'][repoName]['closed'].append(newIssue);

                for issue in openIssues:
                    newIssue = {};
                    newIssue[issue[0]] = issue[1]
                    authors[i]['contributions'][repoName]['open'].append(newIssue);
        helper.writeln('')

    def _ProcessRietveld(self, author, guid):
        reitveldURL = 'https://codereview.chromium.org/search?closed=%s&owner=%s&repo_guid=%s&modified_after=%s&modified_before=%s&limit=30'
        assert(author != None and guid != None)
        email = ''
        if type(author['email']) == list:
            email = author['email'][0]
        else:
            email = author['email']
        closedIssuesURL = reitveldURL % ('2', email, guid, self._current_week[0], self._current_week[1])
        page = requests.get(closedIssuesURL)
        tree = html.fromstring(page.text)
        issueDetails = tree.xpath('//*[@class="subject"]/a/text()')
        closedIssues = []
        for i in xrange(0, len(issueDetails), 2):
            issueURL = 'https://codereview.chromium.org/' + issueDetails[i]
            issueTitle = issueDetails[i + 1].strip()
            closedIssues.append((issueURL, issueTitle))

        openIssuesURL = reitveldURL % ('3', email, guid, self._current_week[0], self._current_week[1])
        page = requests.get(openIssuesURL)
        tree = html.fromstring(page.text)
        issueDetails = tree.xpath('//*[@class="subject"]/a/text()')
        openIssues = []
        for i in xrange(0, len(issueDetails), 2):
            issueURL = 'https://codereview.chromium.org/' + issueDetails[i]
            issueTitle = issueDetails[i + 1].strip()
            openIssues.append((issueURL, issueTitle))

        return (closedIssues, openIssues)

    def _ProcessGithub(self, author, url):
        assert(author['username'] != None)
        githubURL = url + '/pulls?q=is:pr+author:%s+is:%s+updated:%s..%s'
        closedIssuesURL = githubURL % (author['username'], 'closed', self._current_week[0], self._current_week[1])
        page = requests.get(closedIssuesURL)
        tree = html.fromstring(page.text)
        issueDetails = tree.xpath('//*[@class="issue-title-link js-navigation-open"]')
        closedIssues = []
        for issue in issueDetails:
            issueURL = 'https://github.com' + issue.get('href')
            issueTitle = issue.text.strip()
            closedIssues.append((issueURL, issueTitle))

        openIssuesURL = githubURL % (author['username'], 'open', self._current_week[0], self._current_week[1])
        page = requests.get(openIssuesURL)
        tree = html.fromstring(page.text)
        openIssues = []
        issueDetails = tree.xpath('//*[@class="issue-title-link js-navigation-open"]')
        for issue in issueDetails:
            issueURL = 'https://github.com' + issue.get('href')
            issueTitle = issue.text.strip()
            openIssues.append((issueURL, issueTitle))

        return (closedIssues, openIssues)


helper = UtilityHelper(30)
def main():
    (weekStart, weekEnd) = helper.GetCurrentWeek()
    helper.writeln('Generating report for the week: %s - %s\n' % (weekStart, weekEnd))
    # updateRepositories()

    authorContributions = AuthorContributions(helper.GetFormattedAuthorEmails())
    authorContributions.PopulateContributionsAndFormatToJson()

    workbook = ExcelWorkbook()
    workbook.GenerateCompleteContributionsReport()

    helper.writeln('Wow!!! You guys are just awesome, kept me busy till now! See you next week!\n')

if __name__ == '__main__':
    sys.exit(main())
